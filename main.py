import pandas as pd
import json
from moodle_sync import MoodleSync
import openpyxl
from openpyxl.utils import get_column_letter
import random

if __name__ == "__main__":
    import warnings

    warnings.simplefilter("ignore")

    filename = "/Users/dominik/Dropbox/TGM/Systemtechnik SYT/Schuljahr 21_22/20220611_Noten.xlsx"
    course_id = 1086
    datestring = "20220615"
    role_id = 5  # student
    file = openpyxl.load_workbook(filename, data_only=True)
    for i, sheet in enumerate(file.sheetnames):
        print(f"[{i + 1}] {sheet.title()}")

    try:
        choice = int(input("Choose a sheet: ")) - 1
    except ValueError:
        print("Invalid input")
        exit()

    if choice < 0 or choice > len(file.sheetnames):
        print("Invalid input")
        exit()

    # iter over first row to find the last interesting column
    ws = file[file.sheetnames[choice]]
    end_column = ws.max_column + 1
    end_column_letter = get_column_letter(ws.max_column)
    email_column = None
    for cell in ws[1]:
        if cell.value == "Email":
            email_column = cell.column
        if cell.value is None:
            end_column = cell.column
            end_column_letter = cell.column_letter

            break

    # iter over first row to find the last interesting row
    end_row = ws.max_row + 1
    for cell in ws["A"]:
        if cell.value is None:
            end_row = cell.row
            break

    # df = pd.DataFrame(ws.values)
    # print(df)

    from itertools import islice

    data = ws.values
    cols = next(data)[0:end_column - 1]
    data = list(data)
    data = (islice(r, 0, end_column - 1) for r in data[:end_row - 2])
    df = pd.DataFrame(data, columns=cols)
    print(len(df), " Students found")

    for i, col in enumerate(df.columns):
        print(f"[{i + 1}] {col}")

    try:
        choice = int(input("Choose a column with negative competences (ie: 1.1;1.2;): ")) - 1
    except ValueError:
        print("Invalid input")
        exit()

    if choice < 0 or choice > len(df.columns):
        print("Invalid input")
        exit()

    choice_column_name = df.columns[choice]
    df = df[df[choice_column_name].notnull()]

    with open("data/credentials.json", "r") as f:
        credentials = json.load(f)
    url = credentials["url"]
    username = credentials["user"]
    password = credentials["password"]
    service = credentials["service"]
    ms = MoodleSync(url, username, password, service)
    print("Loading student infos from Moodle, please wait...")
    student_info_df = ms.get_enrolled_students(course_id)

    df2 = df.merge(student_info_df, left_on="Email", right_on="email", how="left")

    # show students how are not enrolled
    not_enrolled_df = df2[df2["id"].isnull()]
    print("\nStudents not enrolled:")
    for i, row in not_enrolled_df.iterrows():
        print(f"\t{row['Schüler']} {row['Email']}")
    # ask if you want to enroll them automatically or continue
    choice = input("Do you want to enroll them automatically? (y/n): ")
    if choice == "y":
        # if yes, find students by name? or by email? and enroll them
        emails = not_enrolled_df["Email"].tolist()
        response = ms.get_user_by_email(emails)

        enrolments = []
        for s in response:
            enrolments.append({'roleid': role_id, 'userid': s['id'], 'courseid': course_id})
        print("\nEnrolling students:")
        print("\t")
        print("\n\t".join(f"{s['fullname']} {s['email']}" for s in response))
        # do you want to continue? or stop to enroll them manually
        choice = input(f"Do you want to enroll {len(enrolments)}? (y/n): ")
        if choice == "y":
            r = ms.enroll_students(enrolments)
            # TODO print # of successful enrolments

        choice = input("Do you want to continue? (y/n): ")
        if choice == "y":
            print("Loading student infos from Moodle, please wait...")
            student_info_df = ms.get_enrolled_students(course_id)
            df2 = df.merge(student_info_df, left_on="Email", right_on="email", how="left")
        else:
            exit()
    df2 = df2[df2["id"].notnull()]
    # if no continue
    # iter over chosen column and find all needed competences
    needed_competences = []
    for i, value in df2[choice_column_name].iteritems():
        for competence in value.split(";")[:-1]:
            if competence not in needed_competences:
                needed_competences.append(competence)

    needed_competences = sorted(needed_competences)
    print(";".join(needed_competences), " needed competences found.")

    # create groups
    groups = []
    group_ids = {}
    for c in needed_competences:
        group_id = datestring + str(random.randrange(100, 999))
        group_ids[c] = group_id
        groups.append(
            {"courseid": course_id, "name": f'GruppeSYT{c}@{datestring}', "description": "", "idnumber": group_id})

    print("\nCreating groups:")
    print("\t" + "\n\t".join(f"{g['name']} ({g['idnumber']})" for g in groups))
    choice = input("Do you want to continue? (y/n): ")
    if choice == "y":
        response = ms.create_group(groups)

        groupsids = {}
        for g in response:
            groupsids[g["idnumber"]] = g["id"]

        # add students to groups
        members = []
        for c in needed_competences:
            user_ids = df2[df2[choice_column_name].str.contains(c)]["id"].tolist()
            for user_id in user_ids:
                members.append({"groupid": groupsids[group_ids[c]], "userid": user_id})

        choice = input(f"Do you want to add {len(members)} students to {len(needed_competences)} competences? (y/n): ")
        if choice == "y":
            response = ms.add_students_to_group(members)
            print("Done")

    # create quizes
