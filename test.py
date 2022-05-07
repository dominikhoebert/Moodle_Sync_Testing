import openpyxl

wb = openpyxl.load_workbook('data/20220507_Noten_SYT Grundkompetenztestung.xlsx', data_only=True)
ws = wb["SYT 1abHIT 2122++"]
cell = "AW2"
print(cell, ws[cell].value, ws[cell].data_type)
