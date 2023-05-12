import pandas as pd
import json
import openpyxl
from openpyxl.utils import get_column_letter
import random
from itertools import islice
from datetime import datetime
from moodle_sync_testing import MoodleSyncTesting

role_id = 5  # student


def dialog(file):
    if file.endswith(".xlsx"):
        df = read_in_excel(file)
    elif file.endswith(".csv"):
        df = pd.read_csv(file)
    else:
        print("Error: File must be .xlsx or .csv. Exiting...")
        exit()
    print(len(df), " Students found")

    ms = login_moodle_sync(df)
    ms.course_id = get_course_id(ms)

    for i, col in enumerate(df.columns):
        print(f"[{i + 1}] {col}")

    ms.group_column_name = df.columns[get_column_choice("Groupname Column", df)]
    ms.column_name = df.columns[get_column_choice("MoodleID/Email Column", df)]

    print("Loading student infos from Moodle, please wait...")
    ms.join_enrolled_students()

    choice = input(f"Do you want to enroll missing Students? (y/n): ")
    if choice == "y":
        ms.enroll_students_for_groups()
        ms.join_enrolled_students()

    ms.clean_students()
    ms.log_count_students_in_groups()

    choice = input(f"Do you want to create these groups? (y/n): ")
    if choice != "y":
        print("Exiting...")
        exit()

    ms.create_groups()
    ms.log_groups()

    choice = input(f"Do you want to add this students to the groups? (y/n): ")
    if choice != "y":
        print("Exiting...")
        exit()

    ms.add_students_to_group()
    print("Done")


def get_moodle_id_column_name(df):
    moodle_id_column_name = None
    try:
        choice = int(input("Choose Moodle ID Column (or Q for Email Column instead): ")) - 1
        if choice < 0 or choice > len(df.columns):
            print("Invalid input")
            exit()
        moodle_id_column_name = df.columns[choice]
    except ValueError:
        pass
    return moodle_id_column_name


def get_column_choice(choose_text_string, df):
    try:
        choice = int(input(f"Choose {choose_text_string}: ")) - 1
    except ValueError:
        print("Invalid input")
        exit()
    if choice < 0 or choice > len(df.columns):
        print("Invalid input")
        exit()
    return choice


def get_course_id(ms):
    courses = ms.get_recent_courses()
    for i, course in enumerate(courses.keys()):
        print(f"[{i + 1}] {course}")
    try:
        choice = int(input("Choose a course: ")) - 1
    except ValueError:
        print("Invalid input")
        exit()
    if choice < 0 or choice > len(courses.keys()):
        print("Invalid input")
        exit()
    course_id = list(courses.values())[choice]["id"]
    return course_id


def login_moodle_sync(df):
    credentials_file = input("Enter path to credentials file or 'y' for data/credentials.json: ")
    if credentials_file == "y":
        credentials_file = "./data/credentials.json"
    try:
        with open(credentials_file, "r") as f:
            credentials = json.load(f)
    except FileNotFoundError:
        print(f"Error: {credentials_file} not found. Exiting...")
        exit()
    try:
        url = credentials["url"]
        username = credentials["user"]
        password = credentials["password"]
        service = credentials["service"]
    except KeyError:
        print("Error: Invalid credentials file. Exiting...")
        exit()
    try:
        ms = MoodleSyncTesting(url, username, password, service, None, df, None, None)
    except KeyError:
        print("Error: Failed to connect to Server. Exiting...")
        exit()
    return ms


def read_in_excel(file):
    try:
        file = openpyxl.load_workbook(file, data_only=True)
    except FileNotFoundError:
        print(f"Error: {file} not found. Exiting...")
        exit()
    for i, sheet in enumerate(file.sheetnames):
        print(f"[{i + 1}] {sheet.title()}")
    try:
        choice = int(input("Choose a sheet: ")) - 1
    except ValueError:
        print("Invalid input")
        exit()
    if choice < 0 or choice > len(file.sheetnames):
        print("Invalid input")
        exit()
    ws = file[file.sheetnames[choice]]
    # iter over first row to find the last interesting column
    # find last column with data
    end_column = ws.max_column + 1
    end_column_letter = get_column_letter(ws.max_column)
    for cell in ws[1]:
        if cell.value is None:
            end_column = cell.column
            end_column_letter = cell.column_letter
            break
    # iter over first row to find the last interesting row
    # find last row with data
    end_row = ws.max_row + 1
    for cell in ws["A"]:
        if cell.value is None:
            end_row = cell.row
            break
    data = ws.values
    cols = next(data)[0:end_column - 1]
    data = list(data)
    data = (islice(r, 0, end_column - 1) for r in data[:end_row - 2])
    df = pd.DataFrame(data, columns=cols)
    return df
